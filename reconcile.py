import io
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

POC_CLIENTS = ['Gametime', 'TickPick', 'SeatGeek']
POD_CLIENTS = ['Vivid Seats', 'GoTickets', 'TicketNetwork', 'TicketsNow']
POC_ORDER   = ['Gametime', 'TickPick', 'SeatGeek']
POD_ORDER   = ['Vivid Seats', 'GoTickets', 'TicketNetwork', 'TicketsNow']
POD_RENAME  = {'TicketsNow': 'TicketMaster'}
ALL_SIX     = POC_CLIENTS + POD_CLIENTS

DARK_BLUE = '1F4E79'
MID_BLUE  = '2E75B6'
LT_BLUE   = 'EBF3FB'
WHITE     = 'FFFFFF'
CURR      = '$#,##0.00'

# Human-readable column headers for Part 1
COL_DISPLAY = {
    'PaymentDate':  'Payment Date',
    'DayOfWeek':    'Day of Week',
    'EventName':    'Event Name',
    'EventDate':    'Event Date',
    'TransactionID':'Transaction ID',
    'Description':  'Description',
    'Total':        'Total',
    'InInvoice':    'Found in Y&S',
}


def _thin_border():
    t = Side(style='thin', color='CCCCCC')
    return Border(left=t, right=t, top=t, bottom=t)


def _sc(ws, r, c, val=None, bg=WHITE, fg='000000', bold=False, sz=12,
        ha='left', fmt=None, wrap=False, merge_to=None):
    bdr = _thin_border()
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = Font(name='Calibri', bold=bold, color=fg, size=sz)
    cell.fill      = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal=ha, vertical='center', wrap_text=wrap)
    cell.border    = bdr
    if fmt:
        cell.number_format = fmt
    if merge_to:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge_to)
        for cc in range(c + 1, merge_to + 1):
            ws.cell(row=r, column=cc).fill   = PatternFill('solid', start_color=bg)
            ws.cell(row=r, column=cc).border = bdr
    return cell


# ── Part 1 ────────────────────────────────────────────────────────────────────

def read_csv_utf16(file_bytes: bytes) -> pd.DataFrame:
    text = file_bytes.decode('utf-16-le')
    return pd.read_csv(io.StringIO(text))


def compute_file_total(df: pd.DataFrame) -> float:
    total = 0.0
    for col in ['Proceeds', 'Charges', 'Credit']:
        total += pd.to_numeric(df[col], errors='coerce').fillna(0).sum()
    return round(total, 2)


def match_files_to_payments(csv_files: dict, payment_rows: list):
    payment_totals = {
        row['payment_id']: (
            round(row['proceeds'] + row['charges'] + row['credit'], 2),
            row['date'],
        )
        for row in payment_rows
    }
    file_totals = {}
    for fname, fbytes in csv_files.items():
        try:
            df = read_csv_utf16(fbytes)
            file_totals[fname] = compute_file_total(df)
        except Exception as e:
            file_totals[fname] = None
            print(f'Warning: could not read {fname}: {e}')

    matched = {}
    unmatched_files = []
    for fname, ftotal in file_totals.items():
        if ftotal is None:
            unmatched_files.append(fname)
            continue
        found = False
        for pid, (ptotal, pdate) in payment_totals.items():
            if abs(ftotal - ptotal) < 0.05 and pid not in matched:
                matched[pid] = (fname, pdate)
                found = True
                break
        if not found:
            unmatched_files.append(fname)

    unmatched_payments = [pid for pid in payment_totals if pid not in matched]
    return matched, unmatched_files, unmatched_payments


def build_part1_df(csv_files: dict, matched: dict, invoice_bytes: bytes) -> pd.DataFrame:
    inv_df  = pd.read_excel(io.BytesIO(invoice_bytes))
    inv_ids = set(inv_df['Ext Order #'].dropna().astype(str).str.strip())

    dfs = []
    for pid, (fname, date) in matched.items():
        df = read_csv_utf16(csv_files[fname])
        for col in ['Proceeds', 'Charges', 'Credit']:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        df['Total'] = df['Proceeds'] + df['Charges'] + df['Credit']
        dt = pd.to_datetime(date, format='%d/%m/%Y')
        df['PaymentDate'] = dt.strftime('%d/%m/%Y')
        df['DayOfWeek']   = dt.strftime('%A')
        df = df.drop(columns=['Venue', 'Proceeds', 'Charges', 'Credit'], errors='ignore')
        dfs.append(df)

    combined = pd.concat(dfs, ignore_index=True)
    cols = ['PaymentDate', 'DayOfWeek', 'EventName', 'EventDate',
            'TransactionID', 'Description', 'Total']
    extra = [c for c in combined.columns if c not in cols]
    combined = combined[cols + extra]
    def _flag(row, inv_ids):
        if row['Total'] < 0 and str(row.get('Description', '')).strip() in ('', 'nan', 'None'):
            return 'Yes'
        return 'Yes' if str(row['TransactionID']).strip() in inv_ids else 'No'
    combined['InInvoice'] = combined.apply(lambda row: _flag(row, inv_ids), axis=1)
    return combined


# ── Part 2 ────────────────────────────────────────────────────────────────────

def build_part2_data(inv_det_bytes: bytes, inv_bytes: bytes):
    inv_det = pd.read_excel(io.BytesIO(inv_det_bytes))
    inv     = pd.read_excel(io.BytesIO(inv_bytes))

    inv_det = inv_det[inv_det['Client'].isin(ALL_SIX)].copy()
    inv     = inv[inv['Client'].isin(POD_CLIENTS)].copy()

    inv_det['Created Date'] = pd.to_datetime(inv_det['Created Date'])
    date_min   = inv_det['Created Date'].min().strftime('%b %-d, %Y')
    date_max   = inv_det['Created Date'].max().strftime('%b %-d, %Y')
    date_range = f'{date_min} - {date_max}'

    poc_sales = inv_det[inv_det['Client'].isin(POC_CLIENTS)].groupby('Client')['Total Price'].sum()
    poc_rows  = [(cl, round(poc_sales.get(cl, 0) / 1000) * 1000) for cl in POC_ORDER]
    poc_total = sum(r[1] for r in poc_rows)

    pod_det    = inv_det[inv_det['Client'].isin(POD_CLIENTS)].copy()
    inv_orders = set(inv['Ext Order #'].dropna().astype(str))
    pod_det['matched'] = pod_det['Ext Order #'].astype(str).isin(inv_orders)

    new_sales = pod_det.groupby('Client')['Total Price'].sum()
    adj60     = pod_det[~pod_det['matched']].groupby('Client')['Total Price'].sum() * 0.60
    unpaid    = inv.groupby('Client')['Bal.'].sum()

    pod_data = []
    for cl in POD_ORDER:
        display = POD_RENAME.get(cl, cl)
        ns = new_sales.get(cl, 0)
        a6 = adj60.get(cl, 0)
        up = unpaid.get(cl, 0)
        pj = round((a6 + up) / 1000) * 1000
        pod_data.append((display, ns, a6, up, pj))

    return date_range, poc_rows, poc_total, pod_data


# ── Combined output ───────────────────────────────────────────────────────────

def _write_stubhub_pivot(ws, df: pd.DataFrame, start_row: int = 2):
    """Pivot: rows = Found in Y&S / Payment Date, values = Sum of Total."""
    bdr = _thin_border()

    # Build pivot data
    pivot = df.groupby(['InInvoice', 'PaymentDate'])['Total'].sum().reset_index()
    pivot.columns = ['Found in Y&S', 'Payment Date', 'Sum of Total']

    # Sort: Yes first, then No; within each group sort by date
    pivot['_sort'] = pivot['Found in Y&S'].map({'Yes': 0, 'No': 1})
    pivot['_date'] = pd.to_datetime(pivot['Payment Date'], format='%d/%m/%Y')
    pivot = pivot.sort_values(['_sort', '_date']).drop(columns=['_sort', '_date'])

    # Place pivot header in column J onwards (col 10) to avoid overlapping detail
    PCOL = 10  # pivot starts at column J

    # Header
    ws.row_dimensions[start_row].height = 20
    for ci, name in enumerate(['Found in Y&S', 'Payment Date', 'Sum of Total'], 1):
        c = ws.cell(row=start_row, column=PCOL + ci - 1, value=name)
        c.font      = Font(name='Calibri', bold=True, color=WHITE, size=12)
        c.fill      = PatternFill('solid', start_color=DARK_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = bdr

    # Set pivot column widths
    for i, w in enumerate([18, 16, 18], PCOL):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Group totals tracker
    current_group = None
    row = start_row + 1

    rows_list = pivot.values.tolist()
    for i, (found, pdate, total) in enumerate(rows_list):
        ws.row_dimensions[row].height = 16
        # Detect group change
        if found != current_group:
            current_group = found
            group_start = row

        fill_w = PatternFill('solid', start_color=WHITE)

        c1 = ws.cell(row=row, column=PCOL, value=found)
        c1.font = Font(name='Calibri', size=12); c1.fill = fill_w; c1.border = bdr
        c1.alignment = Alignment(horizontal='center', vertical='center')

        c2 = ws.cell(row=row, column=PCOL+1, value=pdate)
        c2.font = Font(name='Calibri', size=12); c2.fill = fill_w; c2.border = bdr
        c2.alignment = Alignment(horizontal='center', vertical='center')

        rounded_total = round(total / 1000) * 1000
        c3 = ws.cell(row=row, column=PCOL+2, value=rounded_total)
        c3.font = Font(name='Calibri', size=12); c3.fill = fill_w; c3.border = bdr
        c3.alignment = Alignment(horizontal='center', vertical='center')
        c3.number_format = CURR

        # Check if next row is a different group or end — write subtotal
        next_group = rows_list[i+1][0] if i+1 < len(rows_list) else None
        if next_group != found:
            row += 1
            ws.row_dimensions[row].height = 16
            subtotal = sum(r[2] for r in rows_list if r[0] == found)
            for ci, col in enumerate([PCOL, PCOL+1, PCOL+2], 0):
                c = ws.cell(row=row, column=col)
                c.font   = Font(name='Calibri', bold=True, color=WHITE, size=12)
                c.fill   = PatternFill('solid', start_color=DARK_BLUE)
                c.border = bdr
                c.alignment = Alignment(horizontal='left' if ci == 0 else 'center', vertical='center')
            ws.cell(row=row, column=PCOL).value = f'{found} — Subtotal'
            ws.cell(row=row, column=PCOL+2).value = round(subtotal / 1000) * 1000
            ws.cell(row=row, column=PCOL+2).number_format = CURR
            row += 1  # blank gap between groups
            ws.row_dimensions[row].height = 8
            row += 1

        row += 1

    # Grand total
    ws.row_dimensions[row].height = 16
    grand = df['Total'].sum()
    for ci, col in enumerate([PCOL, PCOL+1, PCOL+2], 0):
        c = ws.cell(row=row, column=col)
        c.font   = Font(name='Calibri', bold=True, color=WHITE, size=12)
        c.fill   = PatternFill('solid', start_color=MID_BLUE)
        c.border = bdr
        c.alignment = Alignment(horizontal='left' if ci == 0 else 'center', vertical='center')
    ws.cell(row=row, column=PCOL).value = 'GRAND TOTAL'
    ws.cell(row=row, column=PCOL+2).value = round(grand / 1000) * 1000
    ws.cell(row=row, column=PCOL+2).number_format = CURR
    return row


def write_combined_xlsx(part1_df: pd.DataFrame, part2_data: tuple) -> tuple:
    """Returns (xlsx_bytes, filename)."""
    date_range, poc_rows, poc_total, pod_data = part2_data

    # Build filename from date range in part2
    # date_range like "May 4, 2026 - May 10, 2026"
    try:
        parts = date_range.split(' - ')
        d1 = pd.to_datetime(parts[0]).strftime('%b %-d')
        d2 = pd.to_datetime(parts[1]).strftime('%b %-d')
        filename = f'Cash Inflow Projection {d1} thru {d2}.xlsx'
    except Exception:
        filename = 'Cash Inflow Projection.xlsx'

    wb = Workbook()

    ws1 = wb.active
    ws1.title = 'StubHub'
    pivot_rows = _write_stubhub_pivot(ws1, part1_df, start_row=1)
    _write_part1_sheet(ws1, part1_df, start_row=pivot_rows + 3)

    ws2 = wb.create_sheet('Other Networks')
    _write_part2_sheet(ws2, date_range, poc_rows, poc_total, pod_data)

    wb.active = ws1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename


def _write_part1_sheet(ws, df: pd.DataFrame, start_row: int = 1):
    bdr = _thin_border()
    hdr_cols = df.columns.tolist()
    display_headers = [COL_DISPLAY.get(c, c) for c in hdr_cols]

    hdr_row = start_row
    for col_idx, display_name in enumerate(display_headers, 1):
        c = ws.cell(row=hdr_row, column=col_idx, value=display_name)
        c.font      = Font(name='Calibri', bold=True, color=WHITE, size=12)
        c.fill      = PatternFill('solid', start_color=DARK_BLUE)
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border    = bdr

    fill_b   = PatternFill('solid', start_color=WHITE)
    fill_yes = PatternFill('solid', start_color='E2EFDA')
    fill_no  = PatternFill('solid', start_color='FCE4D6')
    data_fnt = Font(name='Calibri', size=12)

    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        base_fill = fill_b
        for col_idx, value in enumerate(row, 1):
            c = ws.cell(row=row_idx, column=col_idx, value=value)
            c.font   = data_fnt
            c.border = bdr
            c.alignment = Alignment(vertical='center')
            col_name = hdr_cols[col_idx - 1]
            if col_name == 'Total':
                c.fill          = base_fill
                c.number_format = CURR
            elif col_name == 'InInvoice':
                c.fill      = fill_yes if value == 'Yes' else fill_no
                c.alignment = Alignment(horizontal='center', vertical='center')
            else:
                c.fill = base_fill

    # Column widths — use display header length as minimum
    for col_idx, (col_name, display_name) in enumerate(zip(hdr_cols, display_headers), 1):
        data_max = df.iloc[:, col_idx - 1].astype(str).str.len().max()
        max_len  = max(len(display_name), data_max)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

    ws.row_dimensions[hdr_row].height = 20
    ws.freeze_panes = f'A{start_row + 1}'
    return start_row + len(df)


def _write_part2_sheet(ws, date_range, poc_rows, poc_total, pod_data):
    for i, w in enumerate([22, 18, 16, 32, 16], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 22
    _sc(ws, 1, 1, f'Week of  {date_range}',
        bg=MID_BLUE, fg=WHITE, sz=12, ha='left', merge_to=5)

    ws.row_dimensions[2].height = 24
    _sc(ws, 2, 1, 'Pay on Confirmation',
        bg=DARK_BLUE, fg=WHITE, bold=True, sz=12, merge_to=5)
    ws.row_dimensions[3].height = 13
    _sc(ws, 3, 1, 'Gametime - TickPick - SeatGeek',
        fg='777777', sz=9, merge_to=2)
    ws.row_dimensions[4].height = 18
    _sc(ws, 4, 1, 'Network',   bg=MID_BLUE, fg=WHITE, bold=True, ha='center')
    _sc(ws, 4, 2, 'New Sales', bg=MID_BLUE, fg=WHITE, bold=True, ha='center')

    for ri, (client, sales) in enumerate(poc_rows, 5):
        ws.row_dimensions[ri].height = 16
        _sc(ws, ri, 1, client, bg=WHITE)
        _sc(ws, ri, 2, sales,  bg=WHITE, ha='center', fmt=CURR)

    tr1 = 5 + len(poc_rows)
    ws.row_dimensions[tr1].height = 16
    _sc(ws, tr1, 1, 'TOTAL',    bg=DARK_BLUE, fg=WHITE, bold=True)
    _sc(ws, tr1, 2, poc_total,  bg=DARK_BLUE, fg=WHITE, bold=True, ha='center', fmt=CURR)

    gap = tr1 + 2
    ws.row_dimensions[gap].height = 24
    _sc(ws, gap, 1, 'Pay on Delivery',
        bg=DARK_BLUE, fg=WHITE, bold=True, sz=12, merge_to=5)
    ws.row_dimensions[gap + 1].height = 13
    _sc(ws, gap + 1, 1, 'Vivid Seats - GoTickets - TicketNetwork - TicketMaster',
        fg='777777', sz=9, merge_to=5)

    hdr = gap + 2
    ws.row_dimensions[hdr].height = 30
    for ci, nm in enumerate(['Network', 'New Sales\nfor Future Events', 'Adjusted (60%)',
                              'Unpaid Sales for\nEvents Last Week', 'Projected Inflow'], 1):
        _sc(ws, hdr, ci, nm, bg=MID_BLUE, fg=WHITE, bold=True, ha='center', wrap=True)

    for ri, (cl, ns, a6, up, pj) in enumerate(pod_data, hdr + 1):
        ws.row_dimensions[ri].height = 16
        _sc(ws, ri, 1, cl, bg=WHITE)
        _sc(ws, ri, 2, ns, bg=WHITE, ha='center', fmt=CURR)
        _sc(ws, ri, 3, a6, bg=WHITE, ha='center', fmt=CURR)
        _sc(ws, ri, 4, up, bg=WHITE, ha='center', fmt=CURR)
        _sc(ws, ri, 5, pj, bg=WHITE, ha='center', fmt=CURR)

    tr2 = hdr + 1 + len(pod_data)
    ws.row_dimensions[tr2].height = 16
    _sc(ws, tr2, 1, 'TOTAL',                     bg=DARK_BLUE, fg=WHITE, bold=True)
    _sc(ws, tr2, 2, sum(r[1] for r in pod_data), bg=DARK_BLUE, fg=WHITE, bold=True, ha='center', fmt=CURR)
    _sc(ws, tr2, 3, sum(r[2] for r in pod_data), bg=DARK_BLUE, fg=WHITE, bold=True, ha='center', fmt=CURR)
    _sc(ws, tr2, 4, sum(r[3] for r in pod_data), bg=DARK_BLUE, fg=WHITE, bold=True, ha='center', fmt=CURR)
    _sc(ws, tr2, 5, sum(r[4] for r in pod_data), bg=DARK_BLUE, fg=WHITE, bold=True, ha='center', fmt=CURR)
